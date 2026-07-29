import openpyxl
import json
import os
import sys
import zipfile
from openpyxl.xml.functions import fromstring

xlsx_path = sys.argv[1] if len(sys.argv) > 1 else None
if not xlsx_path:
    print("Usage: python extract_model_features.py <xlsx_path>")
    sys.exit(1)

wb = openpyxl.load_workbook(xlsx_path)
ws = wb[wb.sheetnames[0]]

FEATURE_KEYS = [
    'vllm_ascend_balance_scheduling',
    'enable_cpu_binding',
    'vllm_ascend_enable_nz',
    'enable_prefix_caching',
    'enable_chunked_prefill',
    'vllm_ascend_enable_mlapo',
    'compilation_config.cudagraph_mode=PIECEWISE',
    'compilation_config.cudagraph_mode=FULL_DECODE_ONLY',
    'compilation_config.cudagraph_mode=FULL',
    'speculative_config.method=mtp',
    'speculative_config.method=eagle3',
    'prefill_context_parallel_size',
    'decode_context_parallel_size',
    'enable_expert_parallel',
    'vllm_ascend_enable_fused_mc2',
    'multistream_overlap_shared_expert',
    'enable_shared_expert_dp',
    'dynamic_eplb',
    'data_parallel_size_local',
    'kv_transfer_config.connector=MooncakeConnectorV1',
    'kv_transfer_config.connector=AscendStoreConnector',
    'vllm_ascend_enable_flashcomm1',
    'finegrained_tp_config.lmhead_tensor_parallel_size',
    'weight_prefetch_config',
]

FEATURE_NAMES = [
    '异步调度',
    '绑核',
    'weight_nz',
    'prefix-cache',
    'chunked-prefill',
    'MLAPO',
    'PIECEWISE',
    'FULL_DECODE_ONLY',
    'FULL',
    'mtp',
    'eagle3',
    'pcp',
    'dcp',
    'EP并行',
    'MoE大融合算子',
    '共享专家多流',
    '共享专家dp',
    'EPLB',
    'DPLB',
    'mooncake池化',
    'kvcache池化',
    'flashcomm1',
    'lm_head_dp',
    '权重预取',
]

FEATURE_CATEGORIES = {
    '通用优化': [
        'vllm_ascend_balance_scheduling',
        'enable_cpu_binding',
        'vllm_ascend_enable_nz',
    ],
    'attention': [
        'enable_prefix_caching',
        'enable_chunked_prefill',
        'vllm_ascend_enable_mlapo',
    ],
    '图模式': [
        'compilation_config.cudagraph_mode=PIECEWISE',
        'compilation_config.cudagraph_mode=FULL_DECODE_ONLY',
        'compilation_config.cudagraph_mode=FULL',
    ],
    '投机推理': [
        'speculative_config.method=mtp',
        'speculative_config.method=eagle3',
    ],
    'CP并行': [
        'prefill_context_parallel_size',
        'decode_context_parallel_size',
    ],
    'MoE': [
        'enable_expert_parallel',
        'vllm_ascend_enable_fused_mc2',
        'multistream_overlap_shared_expert',
        'enable_shared_expert_dp',
    ],
    '负载均衡': [
        'dynamic_eplb',
        'data_parallel_size_local',
    ],
    '池化': [
        'kv_transfer_config.connector=MooncakeConnectorV1',
        'kv_transfer_config.connector=AscendStoreConnector',
    ],
    '其它优化点': [
        'vllm_ascend_enable_flashcomm1',
        'finegrained_tp_config.lmhead_tensor_parallel_size',
        'weight_prefetch_config',
    ],
}

# Scenario tag values (mutually exclusive):
#   "general"           - 通用，所有场景适用
#   "high_concurrency"  - 仅高并发场景适用
#   "long_sequence"     - 仅长序列场景适用
#   "pd_disaggregated"  - 仅PD分离场景适用
#
# Extraction priority:
#   L1: xlsx cell fill color (authoritative, from xlsx author)
#       theme7 (#75BD42, green)  → "general"
#       theme6 (#F2BA02, yellow) → "high_concurrency"
#       theme4 (#4874CB, blue)   → "long_sequence"
#   L2: category-based inference (for features without cell color)
#   L3: SCENARIO_OVERRIDES (manual, covers edge cases L1/L2 miss)

# L2: category → scenario inference
CATEGORY_SCENARIO_MAP = {
    '通用优化': 'general',
    'attention': 'general',
    '图模式': 'general',
    '投机推理': 'general',
    'CP并行': 'long_sequence',
    'MoE': 'general',
    '负载均衡': 'high_concurrency',
    '池化': 'pd_disaggregated',
    '其它优化点': 'general',
}

# L3: manual overrides (highest priority, overrides L1 color and L2 category)
# Only add entries where L1/L2 result is incorrect based on semantic analysis
SCENARIO_OVERRIDES = {
    # EPLB: solves expert hot-spot from many concurrent requests;
    # low concurrency has no hot-spot problem, overhead > benefit
    'dynamic_eplb': 'high_concurrency',
    # 共享专家dp: DP for shared experts → high_concurrency
    'enable_shared_expert_dp': 'high_concurrency',
    # lm_head_dp: reduces TP communication → high_concurrency
    'finegrained_tp_config.lmhead_tensor_parallel_size': 'high_concurrency',
    # 权重预取: prefetch weights to NPU → general (useful in all scenarios)
    'weight_prefetch_config': 'general',
}


def parse_cell(val):
    if val is None or str(val).strip() == '-':
        return None
    s = str(val).strip()
    if s == '√':
        return 1
    if s == '×':
        return 0
    if s.startswith('√'):
        return s[1:]
    if s.startswith('×'):
        return '不支持' + s[1:]
    return s


def model_name_to_filename(model_name):
    return model_name.replace('.', '_').replace('/', '_').replace(' ', '_').lower()


def extract_scenario_tags_from_xlsx(ws, feature_keys):
    """Extract scenario tags from xlsx cell fill colors (L1).

    Returns a dict: {feature_key: scenario_tag} for features with colored cells.
    """
    # Read theme colors from xlsx
    theme_colors = []
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            theme_xml = z.read('xl/theme/theme1.xml')
        root = fromstring(theme_xml)
        ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
        clr_scheme = root.find('.//a:clrScheme', ns)
        for tag in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2',
                     'accent3', 'accent4', 'accent5', 'accent6',
                     'hlink', 'folHlink']:
            elem = clr_scheme.find(f'a:{tag}', ns)
            for child in elem:
                if 'val' in child.attrib:
                    theme_colors.append(child.attrib['val'])
                break
    except Exception:
        pass

    # Theme index → scenario tag
    # theme7 (#75BD42) = green  → general
    # theme6 (#F2BA02) = yellow → high_concurrency
    # theme4 (#4874CB) = blue   → long_sequence
    THEME_SCENARIO_MAP = {
        7: 'general',
        6: 'high_concurrency',
        4: 'long_sequence',
    }

    # Scan data rows, collect color per column
    col_scenario = {}
    for row_idx in range(3, ws.max_row + 1):
        for i, key in enumerate(feature_keys):
            if key in col_scenario:
                continue
            col = 5 + i
            cell = ws.cell(row=row_idx, column=col)
            fill = cell.fill
            if fill.patternType and fill.patternType != 'none':
                fg = fill.fgColor
                if fg.type == 'theme' and fg.theme in THEME_SCENARIO_MAP:
                    col_scenario[key] = THEME_SCENARIO_MAP[fg.theme]

    return col_scenario


def build_scenario_tags(feature_keys, feature_categories, ws):
    """Build scenario_tags with L1 (color) → L2 (category) → L3 (override)."""
    # L1: extract from xlsx colors
    color_tags = extract_scenario_tags_from_xlsx(ws, feature_keys)

    # L2: category-based inference
    category_tags = {}
    for cat, keys in feature_categories.items():
        scenario = CATEGORY_SCENARIO_MAP.get(cat, 'general')
        for key in keys:
            category_tags[key] = scenario

    # Merge: L1 > L2 > L3
    scenario_tags = {}
    for key in feature_keys:
        if key in SCENARIO_OVERRIDES:
            scenario_tags[key] = SCENARIO_OVERRIDES[key]
        elif key in color_tags:
            scenario_tags[key] = color_tags[key]
        elif key in category_tags:
            scenario_tags[key] = category_tags[key]
        else:
            scenario_tags[key] = 'general'

    return scenario_tags


common = {
    'features': dict(zip(FEATURE_KEYS, FEATURE_NAMES)),
}

models = {}

current_model = None
current_params = None
current_quant = None

for row_idx in range(3, ws.max_row + 1):
    a_val = ws.cell(row=row_idx, column=1).value
    b_val = ws.cell(row=row_idx, column=2).value
    c_val = ws.cell(row=row_idx, column=3).value
    d_val = ws.cell(row=row_idx, column=4).value

    if a_val and str(a_val).strip() and str(a_val).strip() not in ['备注', '']:
        current_model = str(a_val).strip()
    if b_val and str(b_val).strip():
        current_params = str(b_val).strip()
    if c_val and str(c_val).strip():
        current_quant = str(c_val).strip()

    if not d_val or not str(d_val).strip():
        continue

    version = str(d_val).strip()
    feat_vals = {}
    for i, key in enumerate(FEATURE_KEYS):
        cell_val = ws.cell(row=row_idx, column=5 + i).value
        parsed = parse_cell(cell_val)
        if parsed is not None:
            feat_vals[key] = parsed

    if current_model and feat_vals:
        if current_model not in models:
            models[current_model] = {}
        if current_params not in models[current_model]:
            models[current_model][current_params] = {}
        if current_quant not in models[current_model][current_params]:
            models[current_model][current_params][current_quant] = {}
        models[current_model][current_params][current_quant][version] = feat_vals

# Build scenario tags and group into two-level hierarchy
scenario_tags = build_scenario_tags(FEATURE_KEYS, FEATURE_CATEGORIES, ws)

# Two-level grouping:
# Level 1: general_configs (混部/PD分离均可) | pd_disaggregated_only (仅PD分离)
# Level 2 (under general_configs): general | long_sequence | high_concurrency
scenario_groups = {
    'general_configs': {
        'general': [],
        'long_sequence': [],
        'high_concurrency': [],
    },
    'pd_disaggregated_only': [],
}
for key in FEATURE_KEYS:
    tag = scenario_tags[key]
    if tag == 'pd_disaggregated':
        scenario_groups['pd_disaggregated_only'].append(key)
    elif tag in scenario_groups['general_configs']:
        scenario_groups['general_configs'][tag].append(key)
    else:
        scenario_groups['general_configs']['general'].append(key)

common['scenario_groups'] = scenario_groups

# Scenario thresholds: quantitative boundaries for each non-general scenario
common['scenario_thresholds'] = {
    'high_concurrency': {
        'min_concurrency': 16,
        'description': '并发请求数 ≥16 时判定为高并发场景',
    },
    'long_sequence': {
        'min_seq_length': 65536,
        'description': '总上下文长度（输入+输出）≥64K tokens 时判定为长序列场景',
    },
}

script_dir = os.path.dirname(os.path.abspath(__file__))
output_dir = os.path.join(script_dir, '..', 'output')
os.makedirs(output_dir, exist_ok=True)

common_path = os.path.join(output_dir, 'common.json')
with open(common_path, 'w', encoding='utf-8') as f:
    json.dump(common, f, ensure_ascii=False, indent=2)
print(f'Saved: {common_path}')
print(f'Size: {os.path.getsize(common_path)} bytes')

for model_name, model_data in models.items():
    filename = model_name_to_filename(model_name)
    model_path = os.path.join(output_dir, f'model_feature_{filename}.json')
    with open(model_path, 'w', encoding='utf-8') as f:
        json.dump(model_data, f, ensure_ascii=False, indent=2)
    print(f'Saved: {model_path}')
    print(f'Size: {os.path.getsize(model_path)} bytes')
